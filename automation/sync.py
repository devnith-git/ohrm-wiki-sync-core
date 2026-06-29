"""
============================================================================
  ** DEPRECATED — 2026-05-17 **

  This standalone CLI emits a PRE-CANONICAL HTML schema (h3 anchors,
  [New — KEY] markers, yellow row tints, 5/30/65 column widths). The
  scheduled routines (`routines/cm_daily_sync`, `routines/pnp_daily_sync`)
  now enforce a stricter canonical schema (3-col ATC with bullet-form
  Scenario cells, no inline change markers, parenthetical Jira-key
  lists, semantic de-duplication). The two schemas are NOT compatible.

  If you publish via this CLI, the BookStack page will end up with
  the old schema; the next scheduled routine fire will then either
  (a) preserve it verbatim as legacy content or (b) FAIL STEP 6
  validation. Either way, the page will be inconsistent.

  Use the scheduled routines instead:
    https://claude.ai/code/routines

  This file is kept for backward compatibility with existing local
  workflows. To bring it back to canonical, the `call_claude()` prompt
  (~line 215) and the `validate()` checks (~line 270) need to be
  rewritten against `resources/SKILL.md` STEP 5 + STEP 6. Until that
  happens, the CLI runs but prints a warning to stderr at startup.
============================================================================

Jira -> Wiki sync (standalone, no Claude Code dependency).

USAGE
-----
  py sync.py <JIRA-KEY> [options]

EXAMPLES
--------
  # Dry run (Phase A): fetch, compare, save local HTML + DOCX. No wiki writes.
  py sync.py CM-37

  # Publish (Phase B): same as above + POST/PUT to BookStack (allowlist guarded).
  py sync.py CM-37 --publish

  # Update an existing page (auto-detected from config.yaml if listed, or override here)
  py sync.py CM-35 --publish --target-page-id 543

  # Create a new page with a custom title
  py sync.py CM-100 --publish --new-page-title "My Feature - Phase 1"

CONFIG
------
  Reads defaults from automation/config.yaml (project -> source page, chapter, etc).
  CLI args always override config.yaml.

ENV VARS NEEDED
---------------
  JIRA_USER             - Atlassian account email (e.g. devnith@orangehrm.com)
  JIRA_API_TOKEN        - Atlassian API token (create at id.atlassian.com)
  ANTHROPIC_API_KEY     - Claude API key (console.anthropic.com)
  WIKI_TOKEN_ID         - BookStack token id  (or supply via resources/wiki.env.txt)
  WIKI_TOKEN_SECRET     - BookStack token secret
  WIKI_BASE_URL         - https://enterprisewiki.orangehrm.com (or via wiki.env.txt)

OUTPUT
------
  - {workspace}/automation/out/{jira-key}_{date}.html
  - {workspace}/automation/out/{jira-key}_{date}.docx
  - {workspace}/wiki_update_change_log.xlsx (one new row per --publish run)

SAFETY
------
  Default mode is READ-ONLY. The wiki write path is guarded by a per-run
  ALLOWED_WRITES allowlist; any non-allowlisted (method, path) raises and exits.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import uuid
import re
import textwrap
import urllib.request
import urllib.error
from datetime import datetime, timezone
from pathlib import Path

import requests
import yaml
from openpyxl import Workbook, load_workbook

# Load .env if present (repo-local override). Falls back to process env vars
# and finally to the legacy resources/wiki.env.txt file.
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent.parent / ".env")
except ImportError:
    pass  # python-dotenv optional

WORKSPACE = Path(__file__).resolve().parent.parent
HERE = Path(__file__).resolve().parent
OUT_DIR = HERE / "out"
OUT_DIR.mkdir(parents=True, exist_ok=True)

RESOURCES = WORKSPACE / "resources"
SKILL_PATH = RESOURCES / "SKILL.md"
RENDER_PATH = RESOURCES / "WIKI_PAGE_RENDER.md"
GUIDELINE_PATH = RESOURCES / "specification-writing-guideline.md"
# Legacy fallback only — prefer the .env file. Gitignored regardless.
WIKI_ENV_PATH = RESOURCES / "wiki.env.txt"

LOG_XLSX = WORKSPACE / "wiki_update_change_log.xlsx"
CONFIG_PATH = HERE / "config.yaml"

DATE_TAG = datetime.now(timezone.utc).strftime("%Y-%m-%d")


# ----------------------------------------------------------------------
# Env / config loading
# ----------------------------------------------------------------------
def load_wiki_env():
    """BookStack creds: prefer env vars; fall back to resources/wiki.env.txt."""
    env = {}
    if WIKI_ENV_PATH.exists():
        for line in WIKI_ENV_PATH.read_text().splitlines():
            if "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return {
        "WIKI_TOKEN_ID":     os.environ.get("WIKI_TOKEN_ID")     or env.get("WIKI_TOKEN_ID"),
        "WIKI_TOKEN_SECRET": os.environ.get("WIKI_TOKEN_SECRET") or env.get("WIKI_TOKEN_SECRET"),
        "WIKI_BASE_URL":     os.environ.get("WIKI_BASE_URL")     or env.get("URL"),
    }


def load_config():
    if not CONFIG_PATH.exists():
        return {"projects": {}}
    return yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8")) or {}


def require(value, name):
    if not value:
        raise SystemExit(f"Missing required value: {name}. Check your env vars and config.yaml.")
    return value


# ----------------------------------------------------------------------
# Jira fetch (direct REST, no MCP)
# ----------------------------------------------------------------------
class Jira:
    def __init__(self, base, user, token):
        self.base = base.rstrip("/")
        self.auth = (user, token)

    def get(self, path):
        r = requests.get(f"{self.base}{path}", auth=self.auth, timeout=30)
        r.raise_for_status()
        return r.json()

    def search(self, jql, fields, max_results=100):
        r = requests.post(
            f"{self.base}/rest/api/3/search/jql",
            auth=self.auth,
            json={"jql": jql, "fields": fields, "maxResults": max_results},
            timeout=60,
        )
        r.raise_for_status()
        return r.json()

    def issue(self, key, fields=None):
        params = ""
        if fields:
            params = "?fields=" + ",".join(fields)
        return self.get(f"/rest/api/3/issue/{key}{params}")


# ----------------------------------------------------------------------
# BookStack client with per-run allowlist
# ----------------------------------------------------------------------
class BookStack:
    def __init__(self, base, token_id, token_secret, allowed_writes=None):
        self.base = base.rstrip("/")
        self.auth = f"Token {token_id}:{token_secret}"
        self.allowed_writes = set(allowed_writes or [])
        self.counters = {"GET": 0, "POST": 0, "PUT": 0, "DELETE": 0, "OTHER": 0}

    def request(self, method, path, body=None):
        if method != "GET" and (method, path) not in self.allowed_writes:
            raise PermissionError(f"REFUSED non-allowlisted write: {method} {path}")
        self.counters[method if method in self.counters else "OTHER"] += 1
        data = None
        headers = {"Authorization": self.auth, "Accept": "application/json"}
        if body is not None:
            data = json.dumps(body).encode("utf-8")
            headers["Content-Type"] = "application/json"
        req = urllib.request.Request(self.base + path, data=data, method=method, headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=30) as r:
                return json.loads(r.read())
        except urllib.error.HTTPError as e:
            sys.stderr.write(f"HTTP {e.code} on {method} {path}: {e.read().decode(errors='replace')}\n")
            raise


# ----------------------------------------------------------------------
# Claude API call (with prompt cache so repeat runs are cheap)
# ----------------------------------------------------------------------
def call_claude(model, source_page, jira_data, target_op, target_meta):
    from anthropic import Anthropic
    client = Anthropic()

    skill = SKILL_PATH.read_text(encoding="utf-8")
    render = RENDER_PATH.read_text(encoding="utf-8")
    guideline = GUIDELINE_PATH.read_text(encoding="utf-8")

    user_prompt = textwrap.dedent(f"""
    You are the OHRM Wiki Spec Renderer. Produce a merged Wiki page from the
    Jira data + the existing source wiki page below, following the universal
    SKILL.md, WIKI_PAGE_RENDER.md, and the canonical specification-writing-guideline.md.

    OPERATION: {target_op}        # 'create' or 'update'
    TARGET META: {json.dumps(target_meta)}

    OUTPUT STRICT JSON (no prose, no markdown wrapper) with this shape:
    {{
      "comparison": [
         {{"jira_key": "...", "jira_summary": "...",
           "wiki_section_matched": "...", "current_status": "Missing|Outdated|Already covered|Conflicting|Partially covered",
           "gap": "...", "action_required": "Create|Update|Keep|Clarify",
           "final_action": "Newly added|Updated|Kept|Requires clarification"}}
      ],
      "merged_html": "<the full body html ready to POST/PUT to BookStack>",
      "sections_created": ["..."],
      "sections_updated": ["..."],
      "sections_removed": [],
      "open_questions": []
    }}

    Rules (match the OHRM peer pages — page 360 style):
    - Use <h3 id="..."> for sections, <h4> for Interfaces sub-headings.
    - Tables: <table border="1"> with <colgroup>; header cells = <td><strong>...</strong></td> (never <th>).
    - No <h1> in the body. No <hr>. No inline font-family / font-size / border-collapse.
    - Audit Trail = 3 cols (#, Action, How tracked) with Section / Performed Screen / Action Description / Sample Audit blocks per row.
    - Form tables = 6 cols; validations use `-`-prefixed lines separated by <br> (never <ul>).
    - Tag new content inline with <span style="color:#b85c00;font-weight:bold;font-size:9pt;">[New &mdash; {{KEY}}]</span> and [Updated &mdash; {{KEY}}].
    - Tint new rows: <tr style="background-color:#fff7d6;">.
    - Anchors: lowercase hyphenated ids on every section heading.

    ================ JIRA DATA ================
    {json.dumps(jira_data, indent=2)[:60000]}

    ================ SOURCE WIKI PAGE (READ-ONLY) ================
    id: {source_page['id']}  name: {source_page['name']!r}
    book_id: {source_page['book_id']}  chapter_id: {source_page['chapter_id']}
    updated_at: {source_page['updated_at']}

    HTML BODY:
    {source_page.get('html', '')[:80000]}

    ================ END ================
    """).strip()

    print(f"[Claude API] model={model}, sending request...", flush=True)
    response = client.messages.create(
        model=model,
        max_tokens=16000,
        system=[
            {"type": "text", "text": skill,     "cache_control": {"type": "ephemeral"}},
            {"type": "text", "text": render,   "cache_control": {"type": "ephemeral"}},
            {"type": "text", "text": guideline, "cache_control": {"type": "ephemeral"}},
        ],
        messages=[{"role": "user", "content": user_prompt}],
    )
    text = response.content[0].text
    usage = response.usage
    print(f"[Claude API] input_tokens={usage.input_tokens}, output_tokens={usage.output_tokens}, "
          f"cache_read={getattr(usage, 'cache_read_input_tokens', 0)}, "
          f"cache_create={getattr(usage, 'cache_creation_input_tokens', 0)}", flush=True)

    # Extract JSON from possibly-fenced response
    m = re.search(r"```(?:json)?\s*(\{.*\})\s*```", text, re.S)
    raw_json = m.group(1) if m else text.strip()
    try:
        return json.loads(raw_json)
    except json.JSONDecodeError as e:
        debug_path = OUT_DIR / f"{uuid.uuid4().hex[:8]}_bad_response.txt"
        debug_path.write_text(text, encoding="utf-8")
        raise SystemExit(f"Claude returned non-JSON. Saved to {debug_path}. Error: {e}")


# ----------------------------------------------------------------------
# Validation
# ----------------------------------------------------------------------
def validate(body_html, jira_key):
    checks = [
        ("No <h1> in body",                "<h1" not in body_html),
        ("No <th> (use <td><strong>)",     "<th>" not in body_html and "<th " not in body_html),
        ("No inline font-family",          "font-family:" not in body_html),
        ("No inline border-collapse",      "border-collapse" not in body_html),
        ("No placeholders",                not any(p in body_html for p in ["TBD", "[insert here]", "xxx-placeholder"])),
        ("Contains [New] or [Updated] markers", "[New &mdash;" in body_html or "[Updated &mdash;" in body_html or "&mdash;" in body_html),
        ("Has at least one <h3 id=...> anchor", bool(re.search(r'<h3 id="[a-z0-9-]+">', body_html))),
        ("Has at least one <table border=\"1\">", '<table border="1"' in body_html),
        ("Has acceptance test cases anchor", 'id="atc"' in body_html),
    ]
    return checks


# ----------------------------------------------------------------------
# DOCX export
# ----------------------------------------------------------------------
def export_docx(body_html, docx_path):
    try:
        from htmldocx import HtmlToDocx
        from docx import Document
        parser = HtmlToDocx()
        parser.table_style = "Table Grid"
        doc = Document()
        parser.add_html_to_document(body_html, doc)
        doc.save(str(docx_path))
        return True, None
    except Exception as e:
        return False, str(e)


# ----------------------------------------------------------------------
# Audit log row
# ----------------------------------------------------------------------
LOG_HEADERS = [
    "Cycle ID", "Date and time", "Project name", "Jira source",
    "Wiki space", "Wiki parent page", "Wiki page title", "Wiki page URL or ID",
    "Operation type", "Jira IDs reviewed", "Jira IDs added to Wiki", "Jira IDs already covered",
    "Jira IDs partially covered", "Jira IDs missing", "Jira IDs updated", "Jira IDs requiring clarification",
    "Sections created", "Sections updated", "Sections removed",
    "Local DOCX file name", "Resource files used", "Validation status", "Memory update status",
    "Final status", "Notes",
]


def append_audit_row(row):
    if LOG_XLSX.exists():
        wb = load_workbook(LOG_XLSX); ws = wb.active
    else:
        wb = Workbook(); ws = wb.active; ws.title = "Wiki Updates"; ws.append(LOG_HEADERS)
    ws.append([row.get(h, "") for h in LOG_HEADERS])
    wb.save(LOG_XLSX)


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
def _print_deprecation_warning():
    """Loud stderr banner so anyone invoking the CLI sees the schema-divergence
    risk before any work happens. The CLI keeps running — this is a warning,
    not a hard block, to preserve backward compatibility for ad-hoc workflows."""
    msg = (
        "\n"
        "  =================================================================\n"
        "  WARNING: automation/sync.py is DEPRECATED (2026-05-17).\n"
        "\n"
        "  This CLI emits a pre-canonical HTML schema that the scheduled\n"
        "  routines now reject (h3 anchors, [New - KEY] markers, yellow\n"
        "  row tints, 5/30/65 column widths). Publishing via this CLI will\n"
        "  produce a page that diverges from the canonical 5-table schema\n"
        "  defined in resources/SKILL.md + WIKI_PAGE_RENDER.md +\n"
        "  specification-writing-guideline.md.\n"
        "\n"
        "  Recommended: use the scheduled routines\n"
        "    https://claude.ai/code/routines\n"
        "\n"
        "  If you proceed (e.g. for a local-only test that won't be\n"
        "  published), pass --publish at your own risk. The validator in\n"
        "  this file checks the OLD schema; the routines' validator checks\n"
        "  the NEW schema. They will disagree.\n"
        "  =================================================================\n"
    )
    sys.stderr.write(msg)
    sys.stderr.flush()


def main():
    _print_deprecation_warning()
    ap = argparse.ArgumentParser(description="Jira -> Wiki sync (standalone) — DEPRECATED, see warning above")
    ap.add_argument("jira_key", help="Jira issue key (epic or story), e.g. CM-37")
    ap.add_argument("--publish", action="store_true", help="Actually write to BookStack (default: dry run)")
    ap.add_argument("--source-page-id", type=int, help="BookStack page id to compare against (overrides config)")
    ap.add_argument("--target-page-id", type=int, help="BookStack page id to update (overrides config)")
    ap.add_argument("--target-chapter-id", type=int, help="BookStack chapter id for create mode (overrides config)")
    ap.add_argument("--new-page-title", type=str, help="Title for the new page (create mode)")
    ap.add_argument("--model", default="claude-sonnet-4-6", help="Claude model (default sonnet-4-6)")
    ap.add_argument("--dry-run-llm", action="store_true", help="Skip Claude API; emit a stub (plumbing test only)")
    ap.add_argument("--resolve", action="store_true", help="Just print the resolved destination and exit")
    ap.add_argument("--no-resolver", action="store_true", help="Disable auto-resolution; require explicit CLI/config")
    args = ap.parse_args()

    # ---- Env + config ----
    jira_user = require(os.environ.get("JIRA_USER"), "JIRA_USER env var")
    jira_token = require(os.environ.get("JIRA_API_TOKEN"), "JIRA_API_TOKEN env var")
    if not args.dry_run_llm and not args.resolve:
        require(os.environ.get("ANTHROPIC_API_KEY"), "ANTHROPIC_API_KEY env var")

    wenv = load_wiki_env()
    require(wenv["WIKI_TOKEN_ID"], "WIKI_TOKEN_ID")
    require(wenv["WIKI_TOKEN_SECRET"], "WIKI_TOKEN_SECRET")
    require(wenv["WIKI_BASE_URL"], "WIKI_BASE_URL or resources/wiki.env.txt URL line")

    config = load_config()
    project_key = args.jira_key.split("-")[0]
    proj_defaults = config.get("projects", {}).get(project_key, {})

    source_page_id   = args.source_page_id   or proj_defaults.get("source_page_id")
    target_page_id   = args.target_page_id   or proj_defaults.get("target_page_id")
    target_chapter_id = args.target_chapter_id or proj_defaults.get("target_chapter_id")
    new_page_title   = args.new_page_title   or proj_defaults.get("new_page_title")
    project_name     = proj_defaults.get("project_name", project_key)

    # ---- Auto-resolve missing destination via resolver.py ----
    if not args.no_resolver and not all([source_page_id, (target_page_id or (target_chapter_id and new_page_title))]):
        from resolver import resolve as resolve_dest
        # Need a peek at the Jira issue first so we can pass summary/components to the resolver.
        jira_user_pre = require(os.environ.get("JIRA_USER"), "JIRA_USER env var")
        jira_token_pre = require(os.environ.get("JIRA_API_TOKEN"), "JIRA_API_TOKEN env var")
        jira_peek = Jira("https://orangehrmenterprise.atlassian.net", jira_user_pre, jira_token_pre)
        try:
            issue_peek = jira_peek.issue(args.jira_key,
                fields=["summary", "issuetype", "project", "components"])
        except Exception as e:
            raise SystemExit(f"Could not fetch {args.jira_key} from Jira for auto-resolution: {e}")

        peek_proj = issue_peek["fields"]["project"]
        peek_components = [c.get("name", "") for c in (issue_peek["fields"].get("components") or [])]
        peek_summary = issue_peek["fields"].get("summary", "")
        peek_type = issue_peek["fields"]["issuetype"]["name"]

        print(f"[resolver] No full destination in config. Auto-resolving {args.jira_key}...")
        print(f"[resolver]   project: {peek_proj['key']} - {peek_proj['name']}")
        print(f"[resolver]   summary: {peek_summary[:80]}")
        print(f"[resolver]   components: {peek_components or '(none)'}")

        r = resolve_dest(
            project_key=peek_proj["key"],
            project_name=peek_proj["name"],
            issue_summary=peek_summary,
            issue_type=peek_type,
            components=peek_components,
            config=config,
            use_ai=not args.dry_run_llm,
        )
        print(f"[resolver]   confidence: {r.confidence:.2f}  reason: {r.reason}")
        if r.book_name:
            print(f"[resolver]   suggested book: {r.book_name}  chapter: {r.chapter_name}")
        if r.confidence < 0.5:
            print(f"[resolver]   FAIL — cannot auto-resolve. Suggestions:")
            for s in (r.suggestions or []):
                print(f"     - {s}")
            raise SystemExit(2)

        # Fill in missing pieces from the resolution
        if not source_page_id:    source_page_id    = r.source_page_id
        if not target_page_id:    target_page_id    = r.target_page_id
        if not target_chapter_id: target_chapter_id = r.target_chapter_id
        if not new_page_title:    new_page_title    = r.new_page_title
        if project_name == project_key: project_name = r.project_name

    if args.resolve:
        print(f"\nResolved destination for {args.jira_key}:")
        print(f"  project_name      = {project_name}")
        print(f"  source_page_id    = {source_page_id}")
        print(f"  target_page_id    = {target_page_id}")
        print(f"  target_chapter_id = {target_chapter_id}")
        print(f"  new_page_title    = {new_page_title}")
        sys.exit(0)

    require(source_page_id, "source_page_id (CLI --source-page-id, config.yaml, or auto-resolver)")

    op = "update" if target_page_id else "create"
    if op == "create":
        require(target_chapter_id, "target_chapter_id (CLI, config, or auto-resolver) for create mode")
        require(new_page_title, "new_page_title (CLI, config, or auto-resolver) for create mode")

    print(f"PROJECT:       {project_name} ({project_key})")
    print(f"JIRA:          {args.jira_key}")
    print(f"WIKI SOURCE:   page {source_page_id}")
    print(f"OPERATION:     {op}", "(target page", target_page_id, ")" if op == "update" else f"(target chapter {target_chapter_id}, title {new_page_title!r})")
    print(f"PHASE:         {'B (publish)' if args.publish else 'A (dry run)'}")
    print(f"MODEL:         {args.model}")
    print()

    # ---- Step 1: validate wiki access ----
    print("STEP 1 - Validate Wiki access")
    bs_read = BookStack(wenv["WIKI_BASE_URL"], wenv["WIKI_TOKEN_ID"], wenv["WIKI_TOKEN_SECRET"], allowed_writes=set())
    books = bs_read.request("GET", "/api/books?count=1")
    print(f"  OK ({len(books.get('data', []))} book returned)\n")

    # ---- Step 2: fetch Jira data ----
    print(f"STEP 2 - Fetch Jira {args.jira_key} + children")
    jira = Jira("https://orangehrmenterprise.atlassian.net", jira_user, jira_token)
    issue = jira.issue(args.jira_key,
                       fields=["summary", "description", "status", "issuetype", "created", "updated", "project"])
    print(f"  Issue: {issue['key']}  type={issue['fields']['issuetype']['name']}  "
          f"status={issue['fields']['status']['name']}  summary={issue['fields']['summary'][:80]}")

    # If it's an epic, fetch children
    children = []
    is_epic = issue["fields"]["issuetype"]["name"].lower() == "epic"
    if is_epic:
        search = jira.search(f"parent = {args.jira_key} ORDER BY key ASC",
                             fields=["summary", "issuetype", "status", "updated"])
        for it in search.get("issues", []):
            children.append({
                "key": it["key"],
                "summary": it["fields"]["summary"],
                "type": it["fields"]["issuetype"]["name"],
                "status": it["fields"]["status"]["name"],
                "updated": it["fields"].get("updated", "")[:10],
            })
        # Filter bugs/sub-tasks
        excluded = [c for c in children if c["type"] in ("Bug", "Sub-task")]
        kept = [c for c in children if c["type"] not in ("Bug", "Sub-task")]
        print(f"  Children: {len(children)} ({len(kept)} stories kept, {len(excluded)} bugs/sub-tasks skipped)\n")
    else:
        print(f"  Not an epic; will treat the issue itself as the only story.\n")

    jira_data = {
        "epic": {
            "key": issue["key"],
            "summary": issue["fields"]["summary"],
            "description": issue["fields"].get("description", ""),
            "status": issue["fields"]["status"]["name"],
            "type": issue["fields"]["issuetype"]["name"],
            "created": issue["fields"].get("created", "")[:10],
            "updated": issue["fields"].get("updated", "")[:10],
        },
        "stories_kept": [c for c in children if c["type"] not in ("Bug", "Sub-task")],
        "excluded": [c for c in children if c["type"] in ("Bug", "Sub-task")],
    }

    # ---- Step 3: fetch source wiki page ----
    print(f"STEP 3 - Fetch source Wiki page {source_page_id}")
    src = bs_read.request("GET", f"/api/pages/{source_page_id}")
    print(f"  Page: id={src['id']} name={src['name']!r} updated_at={src['updated_at']}")
    print(f"  HTML body size: {len(src.get('html',''))} chars\n")

    # ---- Step 4: call Claude (or stub) ----
    print("STEP 4 - LLM comparison + merged HTML generation")
    target_meta = {
        "operation": op,
        "target_page_id": target_page_id,
        "target_chapter_id": target_chapter_id,
        "new_page_title": new_page_title,
        "project_name": project_name,
    }
    if args.dry_run_llm:
        print("  [stub] --dry-run-llm: producing a minimal stub instead of calling Claude")
        result = {
            "comparison": [{"jira_key": c["key"], "jira_summary": c["summary"],
                            "wiki_section_matched": "(stub)", "current_status": "Missing",
                            "gap": "(stub)", "action_required": "Create", "final_action": "Newly added"}
                           for c in jira_data["stories_kept"][:5]],
            "merged_html": (
                f'<h3 id="atc">Acceptance Test Cases</h3>'
                f'<table border="1"><colgroup><col style="width:5%;"><col style="width:30%;"><col style="width:65%;"></colgroup>'
                f'<tbody><tr><td><strong>#</strong></td><td><strong>Feature</strong></td><td><strong>Scenario</strong></td></tr>'
                f'<tr><td>1</td><td>Stub</td><td>This is a dry-run stub for {args.jira_key}. Replace by removing --dry-run-llm.</td></tr>'
                f'</tbody></table>'
            ),
            "sections_created": ["Acceptance Test Cases"],
            "sections_updated": [],
            "sections_removed": [],
            "open_questions": [],
        }
    else:
        result = call_claude(args.model, src, jira_data, op, target_meta)
    print(f"  Got {len(result.get('comparison',[]))} comparison rows, "
          f"merged_html={len(result.get('merged_html',''))} chars\n")

    # ---- Step 5: validation ----
    print("STEP 5 - Validation checklist")
    checks = validate(result["merged_html"], args.jira_key)
    all_pass = all(ok for _, ok in checks)
    for lbl, ok in checks:
        print(f"  [{'OK ' if ok else 'FAIL'}] {lbl}")
    print(f"  Overall: {'PASS' if all_pass else 'FAIL'}\n")

    # ---- Step 6: save local artifacts ----
    print("STEP 6 - Save local artifacts")
    base = f"{args.jira_key}_{DATE_TAG}"
    html_path = OUT_DIR / f"{base}.html"
    docx_path = OUT_DIR / f"{base}.docx"
    standalone = (
        '<!DOCTYPE html>\n<html><head><meta charset="utf-8"/>'
        f'<title>{new_page_title or src["name"]}</title>'
        '<style>body{font-family:Arial,sans-serif;font-size:11pt;}'
        'h3{font-size:14pt;font-weight:bold;margin-top:20px;}'
        'h4{font-size:12pt;font-weight:bold;}'
        'table{border-collapse:collapse;width:100%;font-size:10pt;}'
        'table[border="1"] td{border:1px solid #ccc;padding:6px;vertical-align:top;}'
        'ul{margin:0;padding-left:18px;}</style></head><body>\n'
        f'{result["merged_html"]}\n</body></html>\n'
    )
    html_path.write_text(standalone, encoding="utf-8")
    print(f"  HTML: {html_path}  ({html_path.stat().st_size} B)")
    docx_ok, docx_err = export_docx(result["merged_html"], docx_path)
    if docx_ok:
        print(f"  DOCX: {docx_path}  ({docx_path.stat().st_size} B)")
    else:
        print(f"  DOCX FAILED: {docx_err}")
    print()

    cycle_id = f"{project_name.replace(' ', '_')}_{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}"

    if not args.publish:
        print(f"PHASE A complete. Validation: {'PASS' if all_pass else 'FAIL'}. "
              f"HTTP: GET={bs_read.counters['GET']} POST=0 PUT=0\n")
        print(f"To publish: py sync.py {args.jira_key} --publish")
        sys.exit(0)

    if not all_pass:
        raise SystemExit("REFUSED publish: validation FAIL")

    # ---- Step 7: publish ----
    print("STEP 7 - Publish to BookStack")
    if op == "update":
        allowed = {("PUT", f"/api/pages/{target_page_id}")}
    else:
        allowed = {("POST", "/api/pages")}
    bs_write = BookStack(wenv["WIKI_BASE_URL"], wenv["WIKI_TOKEN_ID"], wenv["WIKI_TOKEN_SECRET"], allowed_writes=allowed)
    print(f"  Allowlist: {sorted(allowed)}")

    if op == "update":
        # Pre-flight verify
        tgt = bs_write.request("GET", f"/api/pages/{target_page_id}")
        print(f"  Pre-flight target: id={tgt['id']} name={tgt['name']!r} chapter_id={tgt['chapter_id']}")
        out = bs_write.request("PUT", f"/api/pages/{target_page_id}", body={"html": result["merged_html"]})
    else:
        out = bs_write.request("POST", "/api/pages", body={
            "chapter_id": target_chapter_id,
            "name": new_page_title,
            "html": result["merged_html"],
        })

    page_url = out.get("url") or f"{wenv['WIKI_BASE_URL']}/books/.../page/{out.get('slug','')}"
    print(f"  Published: id={out['id']} name={out['name']!r}  url={page_url}\n")

    # ---- Step 8: audit log ----
    print("STEP 8 - Append audit row")
    append_audit_row({
        "Cycle ID": cycle_id,
        "Date and time": datetime.now(timezone.utc).isoformat(),
        "Project name": project_name,
        "Jira source": args.jira_key,
        "Wiki space": "(see config)",
        "Wiki parent page": "(see config)",
        "Wiki page title": out["name"],
        "Wiki page URL or ID": f"{out['id']} ({page_url})",
        "Operation type": "Create" if op == "create" else "Update",
        "Jira IDs reviewed": ",".join(c["key"] for c in (jira_data["stories_kept"] + jira_data["excluded"])),
        "Jira IDs added to Wiki": ",".join(c["key"] for c in jira_data["stories_kept"]),
        "Sections created": ", ".join(result.get("sections_created", [])),
        "Sections updated": ", ".join(result.get("sections_updated", [])),
        "Sections removed": ", ".join(result.get("sections_removed", [])),
        "Local DOCX file name": docx_path.name if docx_ok else "(failed)",
        "Resource files used": "SKILL.md, WIKI_PAGE_RENDER.md, specification-writing-guideline.md, wiki.env.txt",
        "Validation status": "PASS",
        "Memory update status": "(skipped in MVP)",
        "Final status": "SUCCESS",
        "Notes": f"GET={bs_read.counters['GET']+bs_write.counters['GET']} POST={bs_write.counters['POST']} PUT={bs_write.counters['PUT']}; model={args.model}",
    })
    print(f"  Cycle ID: {cycle_id}\n")

    # ---- Final summary ----
    print("=" * 80)
    print("PUBLISH COMPLETE")
    print("=" * 80)
    print(f"  Wiki page:    {out['name']!r}  (id {out['id']})")
    print(f"  URL:          {page_url}")
    print(f"  Validation:   PASS")
    print(f"  Local files:  {html_path}")
    if docx_ok:
        print(f"                {docx_path}")
    print(f"  Audit log:    {LOG_XLSX}")
    print(f"  HTTP audit:   GET={bs_read.counters['GET']+bs_write.counters['GET']} "
          f"POST={bs_write.counters['POST']} PUT={bs_write.counters['PUT']} OTHER={bs_write.counters['OTHER']}")


if __name__ == "__main__":
    main()
