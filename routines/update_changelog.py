#!/usr/bin/env python3
"""update_changelog.py — append a run's wiki changes to the shared changelog.

CANONICAL STORE = an append-only CSV ledger `logs/changelog/changelog.csv`
(text → never corrupts, merges cleanly across routines/branches). The colored
Excel `logs/changelog/wiki_sync_changelog.xlsx` is a RENDERED VIEW rebuilt from
the full CSV on every run — so the binary is never read-modify-written (the
prior approach base64-corrupted it through the agents' commit flow).

Both files are CANONICAL ON `main`. Every routine, at STEP 9, MUST commit them
to `main` (GitHub Contents API, branch=main) regardless of which branch its
.md audit log lands on — see release-filter-policy.md §17. The routine pulls
the current `main` CSV first, runs this helper, then PUTs both files to `main`.

Usage:
    python routines/update_changelog.py <payload.json> [changelog_dir]

    changelog_dir defaults to logs/changelog

Payload JSON schema (the routine builds this from what it already tracks):
    {
      "run_utc": "2026-06-07T09:05:25Z",
      "run_date_ddmmyyyy": "07/06/2026",   # derived from run_utc if omitted
      "routine": "cm_daily_sync",
      "project": "Compensation Management",
      "status": "NO_CHANGE",
      "changes": [
        {"jira_key","topic","affected_area","wiki_book","wiki_page",
         "wiki_page_id","crud_op","previous_content","new_content",
         "outcome","confidence","fix_version","evidence"}
      ]
    }
A run with no content changes still writes ONE summary row.

Idempotent: rows are de-duplicated by a signature so a retry (e.g. after a
409 on the main PUT) does not double-append. Import-safe: auto-installs
openpyxl; never raises into the routine (prints CHANGELOG-SKIP, exits 0).
"""
import csv
import json
import os
import subprocess
import sys
from datetime import datetime, timezone

CSV_FIELDS = [
    "run_utc", "run_date", "routine", "project", "status", "jira_key",
    "topic", "affected_area", "wiki_book", "wiki_page", "wiki_page_id",
    "crud_op", "previous_content", "new_content", "outcome", "confidence",
    "fix_version", "evidence",
]

PROJECT_COLORS = {
    "Compensation Management": "C6E0B4", "CM": "C6E0B4",
    "PNP": "BDD7EE", "Roster": "FFE699", "Orange Sign": "D9C2EC",
    "CS Features (HT)": "B7DEE8", "CS Features": "B7DEE8",
}
DEFAULT_PROJECT_COLOR = "D9D9D9"
CRUD_COLORS = {
    "created": "70AD47", "updated": "4472C4", "replaced": "ED7D31",
    "migrated": "ED7D31", "deleted": "C00000", "removed": "C00000",
    "no-change": "A6A6A6",
}
STATUS_COLORS = {
    "success": "70AD47", "no_change": "A6A6A6", "no-change": "A6A6A6",
    "skipped": "A6A6A6", "blocked": "FFC000", "failed": "C00000",
}
HEADERS = [
    "Run Time (UTC)", "Routine", "Project", "Status", "Jira Key",
    "Topic / Feature", "Affected Area", "Wiki Book", "Wiki Page", "Page ID",
    "CRUD Op", "Previous Wiki Content", "New / Changed Content", "Outcome",
    "Confidence", "Fix Version", "Evidence / Notes",
]
# CSV field -> xlsx column (run_date is the sheet, not a column)
ROW_ORDER = [
    "run_utc", "routine", "project", "status", "jira_key", "topic",
    "affected_area", "wiki_book", "wiki_page", "wiki_page_id", "crud_op",
    "previous_content", "new_content", "outcome", "confidence",
    "fix_version", "evidence",
]


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        try:
            subprocess.run([sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"], check=True)
            import openpyxl  # noqa: F401
            return True
        except Exception as exc:
            print(f"CHANGELOG-SKIP - openpyxl unavailable and pip install failed: {exc}")
            return False


def _sig(r):
    return "|".join(str(r.get(k, "")) for k in
                     ("run_utc", "routine", "jira_key", "crud_op", "wiki_page", "topic"))


def _load_csv(path):
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def _write_csv(path, rows):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=CSV_FIELDS)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in CSV_FIELDS})


def _ddmmyyyy(run_utc):
    try:
        dt = datetime.fromisoformat(run_utc.replace("Z", "+00:00"))
    except ValueError:
        dt = datetime.now(timezone.utc)
    return dt.strftime("%d/%m/%Y")


def _render_xlsx(rows, xlsx_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    white = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)
    center = Alignment(vertical="center", horizontal="center")
    widths = [17, 22, 18, 10, 11, 26, 18, 20, 22, 8, 12, 40, 40, 14, 11, 11, 30]

    # group rows by run_date (sheet); preserve first-seen order
    sheets = {}
    for r in rows:
        sheets.setdefault(r.get("run_date") or _ddmmyyyy(r.get("run_utc", "")), []).append(r)

    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, srows in sheets.items():
        ws = wb.create_sheet(title=sheet_name.replace("/", ".")[:31])
        for col, head in enumerate(HEADERS, start=1):
            c = ws.cell(1, col, head)
            c.fill, c.font, c.border, c.alignment = header_fill, header_font, border, center
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
        for r in srows:
            row = ws.max_row + 1
            for col, key in enumerate(ROW_ORDER, start=1):
                cell = ws.cell(row, col, r.get(key, ""))
                cell.border, cell.alignment = border, wrap
            proj = r.get("project", "")
            ws.cell(row, 3).fill = PatternFill("solid", fgColor=PROJECT_COLORS.get(proj, DEFAULT_PROJECT_COLOR))
            ws.cell(row, 3).font = Font(bold=True)
            sc = STATUS_COLORS.get(str(r.get("status", "")).lower())
            if sc:
                ws.cell(row, 4).fill = PatternFill("solid", fgColor=sc)
                ws.cell(row, 4).font = white
            cc = CRUD_COLORS.get(str(r.get("crud_op", "")).lower())
            if cc:
                ws.cell(row, 11).fill = PatternFill("solid", fgColor=cc)
                ws.cell(row, 11).font = white
    wb.save(xlsx_path)


def main():
    if len(sys.argv) < 2:
        print("CHANGELOG-SKIP - no payload path given")
        return 0
    payload_path = sys.argv[1]
    cl_dir = sys.argv[2] if len(sys.argv) > 2 else "logs/changelog"
    csv_path = os.path.join(cl_dir, "changelog.csv")
    xlsx_path = os.path.join(cl_dir, "wiki_sync_changelog.xlsx")

    if not _ensure_openpyxl():
        return 0
    try:
        payload = json.load(open(payload_path, encoding="utf-8"))
    except Exception as exc:
        print(f"CHANGELOG-SKIP - cannot read payload {payload_path}: {exc}")
        return 0

    os.makedirs(cl_dir, exist_ok=True)
    run_utc = payload.get("run_utc") or datetime.now(timezone.utc).strftime("%Y-%m-%dT%H%M%SZ")
    run_date = payload.get("run_date_ddmmyyyy") or _ddmmyyyy(run_utc)
    base = {"run_utc": run_utc, "run_date": run_date,
            "routine": payload.get("routine", ""), "project": payload.get("project", ""),
            "status": payload.get("status", "")}
    changes = payload.get("changes") or []
    new_rows = []
    if changes:
        for ch in changes:
            r = dict(base)
            r.update({
                "jira_key": ch.get("jira_key", ""), "topic": ch.get("topic", ""),
                "affected_area": ch.get("affected_area", ""), "wiki_book": ch.get("wiki_book", ""),
                "wiki_page": ch.get("wiki_page", ""), "wiki_page_id": ch.get("wiki_page_id", ""),
                "crud_op": ch.get("crud_op", ""), "previous_content": ch.get("previous_content", "—"),
                "new_content": ch.get("new_content", "—"), "outcome": ch.get("outcome", ""),
                "confidence": ch.get("confidence", ""), "fix_version": ch.get("fix_version", ""),
                "evidence": ch.get("evidence", ""),
            })
            new_rows.append(r)
    else:
        r = dict(base)
        r.update({"topic": "(no changes this run)", "crud_op": "No-change",
                  "previous_content": "—", "new_content": "—",
                  "outcome": base["status"]})
        new_rows.append(r)

    rows = _load_csv(csv_path)
    seen = {_sig(r) for r in rows}
    added = 0
    for r in new_rows:
        if _sig(r) not in seen:
            rows.append(r)
            seen.add(_sig(r))
            added += 1
    _write_csv(csv_path, rows)
    try:
        _render_xlsx(rows, xlsx_path)
        print(f"CHANGELOG-OK - {added} row(s) appended (ledger now {len(rows)} rows); "
              f"sheet '{run_date}'; CSV + xlsx rendered in {cl_dir}")
    except Exception as exc:
        print(f"CHANGELOG-SKIP - CSV updated but xlsx render failed: {exc}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
